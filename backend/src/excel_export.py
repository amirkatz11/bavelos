"""Excel export for CFO-ready discount opportunity reports."""

import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# Note: analyze_invoices and print_summary are imported in main() function only


# Style constants
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(size=16, bold=True, color="FFFFFF")
TITLE_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
LABEL_FONT = Font(size=11, bold=True)
VALUE_FONT = Font(size=11)
HIGH_PRIORITY_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
CURRENCY_FORMAT = "$#,##0.00"
PERCENT_FORMAT = "0.0%"


def export_to_excel(results, output_path="outputs/discount_analysis.xlsx"):
    """Create 3-sheet Excel report from analysis results."""
    
    wb = Workbook()
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    # Create sheets
    create_full_analysis_sheet(wb, results)
    create_top_opportunities_sheet(wb, results)
    create_executive_summary_sheet(wb, results)
    
    # Save
    wb.save(output_path)
    print(f"\n📄 Excel report saved to: {output_path}")
    return output_path


# Alias for compatibility
create_excel_report = export_to_excel


def create_full_analysis_sheet(wb, results):
    """Sheet 1: All invoices with complete data."""
    ws = wb.create_sheet("Full Analysis", 0)
    
    # Headers
    headers = [
        "Invoice Number",
        "Supplier Name",
        "Invoice Amount",
        "Payment Terms",
        "Due Date",
        "Has Discount?",
        "Discount %",
        "Discount Days",
        "Net Days",
        "APR",
        "Net Benefit",
        "Action",
        "Status",
        "Priority",
        "Savings",
        "Reason",
    ]
    
    # Write headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Write data
    for row_idx, result in enumerate(results, 2):
        # Handle both flat structure (new) and nested structure (legacy)
        has_discount = result.get("has_discount") == "Yes" if isinstance(result.get("has_discount"), str) else result.get("has_discount", False)
        
        ws.cell(row=row_idx, column=1, value=result.get("invoice_number", ""))
        ws.cell(row=row_idx, column=2, value=result.get("supplier_name", ""))
        
        # Invoice Amount (currency)
        invoice_amount = result.get("invoice_amount", 0)
        cell = ws.cell(row=row_idx, column=3, value=invoice_amount)
        cell.number_format = CURRENCY_FORMAT
        
        ws.cell(row=row_idx, column=4, value=result.get("payment_terms", ""))
        ws.cell(row=row_idx, column=5, value=result.get("due_date", ""))
        
        # Has Discount?
        ws.cell(row=row_idx, column=6, value=result.get("has_discount", "No"))
        
        # Discount fields
        if has_discount:
            ws.cell(row=row_idx, column=7, value=result.get("discount_percentage", ""))
            ws.cell(row=row_idx, column=8, value=result.get("discount_days", ""))
            ws.cell(row=row_idx, column=9, value=result.get("net_days", ""))
            
            # APR (percent) - convert from percentage to decimal if needed
            apr = result.get("apr", 0)
            if apr and apr > 1:  # If already a percentage (e.g., 36.5), convert to decimal
                apr = apr / 100
            cell = ws.cell(row=row_idx, column=10, value=apr)
            cell.number_format = PERCENT_FORMAT
            
            # Net Benefit (currency)
            net_benefit = result.get("net_benefit", 0)
            cell = ws.cell(row=row_idx, column=11, value=net_benefit)
            cell.number_format = CURRENCY_FORMAT
        else:
            ws.cell(row=row_idx, column=7, value="")
            ws.cell(row=row_idx, column=8, value="")
            ws.cell(row=row_idx, column=9, value=result.get("net_days", ""))
            ws.cell(row=row_idx, column=10, value="")
            ws.cell(row=row_idx, column=11, value="")
        
        # Recommendation fields
        ws.cell(row=row_idx, column=12, value=result.get("action", ""))
        ws.cell(row=row_idx, column=13, value=result.get("status", ""))
        ws.cell(row=row_idx, column=14, value=result.get("priority", ""))
        
        # Savings (currency)
        savings = result.get("savings", 0)
        cell = ws.cell(row=row_idx, column=15, value=savings)
        cell.number_format = CURRENCY_FORMAT
        
        ws.cell(row=row_idx, column=16, value=result.get("reason", ""))
    
    # Auto-size columns
    for col_idx in range(1, len(headers) + 1):
        max_length = 0
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 50)
    
    # Freeze top row
    ws.freeze_panes = "A2"
    
    # Add auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"


def create_top_opportunities_sheet(wb, results):
    """Sheet 2: Top 20 opportunities sorted by savings."""
    ws = wb.create_sheet("Top Opportunities", 1)
    
    # Filter to TAKE DISCOUNT only (handle both flat and nested structures)
    take_discount = [r for r in results if r.get("action") == "TAKE DISCOUNT"]
    
    # Sort by savings (descending)
    top_ops = sorted(take_discount, key=lambda x: x.get("savings", 0), reverse=True)[:20]
    
    if not top_ops:
        ws.cell(row=1, column=1, value="No discount opportunities found")
        return
    
    # Headers (same as Full Analysis)
    headers = [
        "Invoice Number",
        "Supplier Name",
        "Invoice Amount",
        "Payment Terms",
        "Due Date",
        "Has Discount?",
        "Discount %",
        "Discount Days",
        "Net Days",
        "APR",
        "Net Benefit",
        "Action",
        "Status",
        "Priority",
        "Savings",
        "Reason",
    ]
    
    # Write headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Write data
    for row_idx, result in enumerate(top_ops, 2):
        ws.cell(row=row_idx, column=1, value=result.get("invoice_number", ""))
        ws.cell(row=row_idx, column=2, value=result.get("supplier_name", ""))
        
        # Invoice Amount (currency)
        invoice_amount = result.get("invoice_amount", 0)
        cell = ws.cell(row=row_idx, column=3, value=invoice_amount)
        cell.number_format = CURRENCY_FORMAT
        
        ws.cell(row=row_idx, column=4, value=result.get("payment_terms", ""))
        ws.cell(row=row_idx, column=5, value=result.get("due_date", ""))
        ws.cell(row=row_idx, column=6, value=result.get("has_discount", "Yes"))
        
        # Discount fields
        ws.cell(row=row_idx, column=7, value=result.get("discount_percentage", ""))
        ws.cell(row=row_idx, column=8, value=result.get("discount_days", ""))
        ws.cell(row=row_idx, column=9, value=result.get("net_days", ""))
        
        # APR (percent) - convert from percentage to decimal if needed
        apr = result.get("apr", 0)
        if apr and apr > 1:  # If already a percentage (e.g., 36.5), convert to decimal
            apr = apr / 100
        cell = ws.cell(row=row_idx, column=10, value=apr)
        cell.number_format = PERCENT_FORMAT
        
        # Net Benefit (currency)
        net_benefit = result.get("net_benefit", 0)
        cell = ws.cell(row=row_idx, column=11, value=net_benefit)
        cell.number_format = CURRENCY_FORMAT
        
        # Recommendation fields
        ws.cell(row=row_idx, column=12, value=result.get("action", ""))
        ws.cell(row=row_idx, column=13, value=result.get("status", ""))
        
        # Priority (with conditional formatting)
        priority = result.get("priority", "")
        priority_cell = ws.cell(row=row_idx, column=14, value=priority)
        if priority == "High":
            priority_cell.fill = HIGH_PRIORITY_FILL
        
        # Savings (currency)
        savings = result.get("savings", 0)
        cell = ws.cell(row=row_idx, column=15, value=savings)
        cell.number_format = CURRENCY_FORMAT
        
        ws.cell(row=row_idx, column=16, value=result.get("reason", ""))
        
        # Highlight entire row if High priority
        if priority == "High":
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = HIGH_PRIORITY_FILL
    
    # Auto-size columns
    for col_idx in range(1, len(headers) + 1):
        max_length = 0
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 50)
    
    # Freeze top row
    ws.freeze_panes = "A2"


def create_executive_summary_sheet(wb, results):
    """Sheet 3: Executive summary with key metrics."""
    ws = wb.create_sheet("Executive Summary", 2)
    
    # Title
    ws.merge_cells("A1:B1")
    title_cell = ws.cell(row=1, column=1, value="BAVELOS FINOPS - DISCOUNT OPPORTUNITY FINDER")
    title_cell.font = TITLE_FONT
    title_cell.fill = TITLE_FILL
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30
    
    # Calculate metrics (handle both flat and nested structures)
    total_invoices = len(results)
    with_discounts = sum(1 for r in results if r.get("has_discount") == "Yes" or (isinstance(r.get("has_discount"), bool) and r.get("has_discount")))
    take_discount = sum(1 for r in results if r.get("action") == "TAKE DISCOUNT")
    
    total_savings = sum(
        r.get("savings", 0)
        for r in results
        if r.get("action") == "TAKE DISCOUNT"
    )
    
    avg_savings = total_savings / take_discount if take_discount > 0 else 0
    
    # Find highest single opportunity
    highest = max(
        (r for r in results if r.get("action") == "TAKE DISCOUNT"),
        key=lambda x: x.get("savings", 0),
        default=None
    )
    highest_savings = highest.get("savings", 0) if highest else 0
    highest_invoice = highest.get("invoice_number", "N/A") if highest else "N/A"
    
    # Calculate weighted average ROI
    total_roi_weighted = 0
    total_weight = 0
    for r in results:
        if r.get("action") == "TAKE DISCOUNT" and r.get("roi_vs_capital") is not None:
            roi = r.get("roi_vs_capital", 0)
            weight = r.get("invoice_amount", 0)
            total_roi_weighted += roi * weight
            total_weight += weight
    # Convert to decimal for percentage format (e.g., 21.5% -> 0.215)
    weighted_avg_roi = (total_roi_weighted / total_weight / 100) if total_weight > 0 else 0
    
    # Summary metrics (starting at row 4)
    row = 4
    metrics = [
        ("Total Invoices Analyzed", total_invoices, None),
        ("Invoices with Discount Terms", with_discounts, None),
        ("Recommended Actions", take_discount, None),
        ("Total Potential Savings", total_savings, CURRENCY_FORMAT),
        ("Average Savings per Invoice", avg_savings, CURRENCY_FORMAT),
        ("Highest Single Opportunity", f"{highest_invoice} (${highest_savings:,.2f})", None),
        ("ROI vs Cost of Capital (weighted avg)", weighted_avg_roi, PERCENT_FORMAT),
    ]
    
    for label, value, fmt in metrics:
        ws.cell(row=row, column=1, value=label + ":")
        ws.cell(row=row, column=1).font = LABEL_FONT
        
        value_cell = ws.cell(row=row, column=2, value=value)
        value_cell.font = VALUE_FONT
        if fmt:
            value_cell.number_format = fmt
        
        row += 1
    
    # Top 5 suppliers by savings (starting at row 12)
    row = 12
    ws.cell(row=row, column=1, value="Top 5 Suppliers by Savings Potential")
    ws.cell(row=row, column=1).font = LABEL_FONT
    row += 1
    
    # Calculate supplier totals
    supplier_totals = {}
    for r in results:
        if r.get("action") == "TAKE DISCOUNT":
            supplier = r.get("supplier_name", "")
            savings = r.get("savings", 0)
            supplier_totals[supplier] = supplier_totals.get(supplier, 0) + savings
    
    # Sort and get top 5
    top_suppliers = sorted(supplier_totals.items(), key=lambda x: x[1], reverse=True)[:5]
    
    # Table headers
    ws.cell(row=row, column=1, value="Supplier")
    ws.cell(row=row, column=1).font = HEADER_FONT
    ws.cell(row=row, column=1).fill = HEADER_FILL
    ws.cell(row=row, column=2, value="Total Savings")
    ws.cell(row=row, column=2).font = HEADER_FONT
    ws.cell(row=row, column=2).fill = HEADER_FILL
    row += 1
    
    # Table data
    for supplier, savings in top_suppliers:
        ws.cell(row=row, column=1, value=supplier)
        cell = ws.cell(row=row, column=2, value=savings)
        cell.number_format = CURRENCY_FORMAT
        row += 1
    
    # Set column widths
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 25


def main():
    """CLI entry point for Excel export."""
    # Import here to avoid import errors when used as module
    from main import analyze_invoices, print_summary
    
    if len(sys.argv) < 2:
        print("Usage: python src/excel_export.py <path_to_csv>")
        print("Example: python src/excel_export.py data/sample_invoices.csv")
        sys.exit(1)
    
    csv_path = sys.argv[1]
    
    print(f"📊 Analyzing invoices from: {csv_path}")
    results = analyze_invoices(csv_path)
    
    if not results:
        print("⚠️  No results to export.")
        return
    
    output_path = export_to_excel(results)
    print(f"✅ Excel export complete!")


if __name__ == "__main__":
    main()
